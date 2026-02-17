#!/usr/bin/env python3
"""
Batch extraction of metadata from ArcGIS XML (ISO 19139) metadata export files into Excel report.

ArcGIS Pro and ArcGIS Online both can save out metadata in XML format, firstly in an internal 
form based on ISO19139 but including a lot of Esri-specific elements. This script will process 
the files and provide a summary statement of the compliance of the metadata with the ISO19139 
standard, as well as a digest of all the metadata values. The tool also outputs all the metadata 
from each of the XML files processed into one spreadsheet output for ease of access. The tool can 
therefore be used as part of a workflow to successively check for compliance of metadata with the 
ISO19139 standard.

Where 'lookup' codes have been used in the metadata export, the script will resolve them to the 
display label for the code. This is done by looking up the code in the ArcGIS Metadata Details 
Excel file (which is included in the toolkit). The script also outputs a report of the codes 
that have been used in the metadata export, and the display labels for the codes.

The input folder is a script parameter (default: 'xml'). The output report excel file is named
metadata_export_<foldername>.xlsx so it reflects the source folder. Excel reports are placed in a 
folder named 'reports'. The workbook contains three sheets:
  - A "Compliance Summary" sheet: per-file ISO 19139 compliance (Yes/No) and list of
    missing mandatory fields.
  - A "Metadata Export" sheet: one row per file, one column per attribute, with a second
    row labelling each attribute as mandatory, optional, or conditional (ISO 19139/INSPIRE).
  - A "Code Resolution" sheet: a list of all the codes that have been used in the metadata export, 
    and the display labels for the codes.

Progress and errors for each file are printed to the console during the run.

Designed for Esri/ArcGIS ISO 19139-style metadata (e.g. from ArcGIS Online) and aligned
with INSPIRE Regulation 1205/2008 for mandatory/optional/conditional classification.

The script is provided as part of the LandIS Soil Portal project, but can be used with 
any metadata files exported from ArcGIS Pro or ArcGIS Online, including ISO 19139 
metadata files that are properly namespaced.

Stephen Hallett, Cranfield University, 2026

Usage:
    python extract_metadata.py [folder]
    Default folder: xml. Output: metadata_export_<foldername>.xlsx
"""

import argparse
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import html
import re
from pathlib import Path
from collections import OrderedDict

# ISO 19139 / INSPIRE obligation per exported field name (Regulation 1205/2008, INSPIRE TG).
# Used for: (1) the optionality row in "Metadata Export", (2) which fields count as
# mandatory in "Compliance Summary". Keys must match the display names used in
# extract_all_fields(). Unknown fields default to 'optional' in get_field_obligation().
FIELD_OBLIGATION = {
    "Resource Title": "mandatory",
    "Abstract": "mandatory",
    "Topic Category": "mandatory",
    "Keywords": "mandatory",
    "Geographic West Bounding Longitude": "mandatory",
    "Geographic East Bounding Longitude": "mandatory",
    "Geographic North Bounding Latitude": "mandatory",
    "Geographic South Bounding Latitude": "mandatory",
    "Data Language": "mandatory",
    "Scale Denominator": "mandatory",
    "Contact Organisation Name": "mandatory",
    "Contact Email Address": "mandatory",
    "Contact Role": "mandatory",
    "Distribution Online Resource Linkage": "mandatory",
    "Lineage Statement": "mandatory",
    "Data Quality Scope Level": "mandatory",
    "Metadata Language Code": "mandatory",
    "Metadata Date Stamp": "mandatory",
    "Metadata Scope Code": "mandatory",
    "Access Constraints": "mandatory",
    "Conformance Specification Title": "mandatory",
    "Conformance Pass": "mandatory",
    "Publication Date": "conditional",
    "Reference System Code": "conditional",
    "Reference System Code Space": "conditional",
    "Other Constraints": "mandatory",
    "Use Limitation": "mandatory",
}
# All other fields (ArcGIS-specific, optional contact details, etc.) are optional unless listed above.

# ISO 19139 codelist labels for report output (from gmxCodelists etc.).
# ArcGIS exports often use numeric 'value' (e.g. "005" for licence); we map to readable labels.
# by_num (integer -> display label) is sourced from Esri ArcGIS Pro Metadata Toolkit:
#   ArcGIS Metadata Details 20211130.xlsx (Coded Values sheet)
#   https://github.com/Esri/arcgis-pro-metadata-toolkit/blob/master/resources/ArcGIS%20Metadata%20Details%2020211130.xlsx
# See also https://data.noaa.gov/resources/iso19139/schema/resources/Codelist/gmxCodelists.xml
# Keys: code name (lowercase) or 1-based index as string ("1", "005" → 5).

# Excel codelist name -> our _CODELISTS key (when they differ)
_ARCGIS_EXCEL_CODELIST_TO_OURS = {"MD_CharSetCd": "MD_CharacterSetCode"}


def _normalise_code(s):
    """Normalise standard/profile code for by_name lookup (lowercase, no spaces/hyphens)."""
    if not s or not isinstance(s, str):
        return ""
    return re.sub(r"[\s\-/]", "", s).lower().split("(")[0].strip()


def _format_code_as_label(std_code):
    """Turn a camelCase or lowercase code name into a display label (UK spelling where applicable)."""
    if not std_code or not isinstance(std_code, str):
        return str(std_code) if std_code else ""
    s = std_code.strip()
    if s.startswith("(") and "reserved" in s.lower():
        return "Reserved"
    if "/" in s:
        s = s.split("/")[-1]
    out = []
    for i, c in enumerate(s):
        if c.isupper() and i:
            out.append(" ")
            out.append(c)
        else:
            out.append(c.upper() if i == 0 or (i > 0 and out[-1] == " ") else c.lower())
    result = "".join(out).replace("  ", " ").strip()
    if result.lower() == "license":
        return "Licence"
    return result


def _load_arcgis_coded_values_from_excel(excel_path):
    """
    Read ArcGIS Metadata Details xlsx (Coded Values sheet) and return
    list of (our_codelist_name, arc_code_str, std_code_str). Returns None if file missing.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    path = Path(excel_path)
    if not path.is_file():
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["Coded Values"]
        rows = list(ws.iter_rows(min_row=5, max_row=600, values_only=True))
        wb.close()
    except Exception:
        return None
    result = []
    current = None
    for row in rows:
        if row and row[0] is not None:
            current = str(row[0]).strip() if row[0] else current
        if not row or row[2] is None or row[1] is None:
            continue
        codelist_name = _ARCGIS_EXCEL_CODELIST_TO_OURS.get(current, current) if current else None
        if not codelist_name:
            continue
        arc_code = str(row[2]).strip()
        std_code = str(row[1]).strip()
        if arc_code.isdigit():
            result.append((codelist_name, arc_code, std_code))
    return result


def _get_inlined_arcgis_coded_values():
    """
    Inlined (codelist_name, arc_code, std_code) from https://github.com/Esri/arcgis-pro-metadata-toolkit/tree/master/resourcesArcGIS Metadata Details 20211130.xlsx
    so by_num can be built without requiring the Excel file. Used when the file is not present.
    """
    return [
        ("CI_PresentationFormCode", "001", "documentDigital"),
        ("CI_PresentationFormCode", "002", "documentHardcopy"),
        ("CI_PresentationFormCode", "003", "imageDigital"),
        ("CI_PresentationFormCode", "004", "imageHardcopy"),
        ("CI_PresentationFormCode", "005", "mapDigital"),
        ("CI_PresentationFormCode", "006", "mapHardcopy"),
        ("CI_PresentationFormCode", "007", "modelDigital"),
        ("CI_PresentationFormCode", "008", "modelHardcopy"),
        ("CI_PresentationFormCode", "009", "profileDigital"),
        ("CI_PresentationFormCode", "010", "profileHardcopy"),
        ("CI_PresentationFormCode", "011", "tableDigital"),
        ("CI_PresentationFormCode", "012", "tableHardcopy"),
        ("CI_PresentationFormCode", "013", "videoDigital"),
        ("CI_PresentationFormCode", "014", "videoHardcopy"),
        ("CI_PresentationFormCode", "015", "audioDigital"),
        ("CI_PresentationFormCode", "016", "audioHardcopy"),
        ("CI_PresentationFormCode", "017", "multimediaDigital"),
        ("CI_PresentationFormCode", "018", "multimediaHardcopy"),
        ("CI_PresentationFormCode", "019", "diagramDigital"),
        ("CI_PresentationFormCode", "020", "diagramHardcopy"),
        ("CI_PresentationFormCode", "021", "physicalObject"),
        ("CI_RoleCode", "001", "resourceProvider"),
        ("CI_RoleCode", "002", "custodian"),
        ("CI_RoleCode", "003", "owner"),
        ("CI_RoleCode", "004", "user"),
        ("CI_RoleCode", "005", "distributor"),
        ("CI_RoleCode", "006", "originator"),
        ("CI_RoleCode", "007", "pointOfContact"),
        ("CI_RoleCode", "008", "principalInvestigator"),
        ("CI_RoleCode", "009", "processor"),
        ("CI_RoleCode", "010", "publisher"),
        ("CI_RoleCode", "011", "author"),
        ("CI_RoleCode", "012", "collaborator"),
        ("CI_RoleCode", "013", "editor"),
        ("CI_RoleCode", "014", "mediator"),
        ("CI_RoleCode", "015", "rightsHolder"),
        ("CI_RoleCode", "016", "sponsor"),
        ("CI_RoleCode", "017", "coAuthor"),
        ("CI_RoleCode", "018", "contributor"),
        ("CI_RoleCode", "019", "funder"),
        ("CI_RoleCode", "020", "stakeholder"),
        ("MD_MaintenanceFrequencyCode", "001", "continual"),
        ("MD_MaintenanceFrequencyCode", "002", "daily"),
        ("MD_MaintenanceFrequencyCode", "003", "weekly"),
        ("MD_MaintenanceFrequencyCode", "004", "fortnightly"),
        ("MD_MaintenanceFrequencyCode", "005", "monthly"),
        ("MD_MaintenanceFrequencyCode", "006", "quarterly"),
        ("MD_MaintenanceFrequencyCode", "007", "biannually"),
        ("MD_MaintenanceFrequencyCode", "008", "annually"),
        ("MD_MaintenanceFrequencyCode", "009", "asNeeded"),
        ("MD_MaintenanceFrequencyCode", "010", "irregular"),
        ("MD_MaintenanceFrequencyCode", "011", "notPlanned"),
        ("MD_MaintenanceFrequencyCode", "012", "unknown"),
        ("MD_MaintenanceFrequencyCode", "013", "semimonthly"),
        ("MD_MaintenanceFrequencyCode", "014", "periodic"),
        ("MD_MaintenanceFrequencyCode", "015", "biennially"),
        ("MD_ProgressCode", "001", "completed"),
        ("MD_ProgressCode", "002", "historicalArchive"),
        ("MD_ProgressCode", "003", "obsolete"),
        ("MD_ProgressCode", "004", "onGoing"),
        ("MD_ProgressCode", "005", "planned"),
        ("MD_ProgressCode", "006", "required"),
        ("MD_ProgressCode", "007", "underDevelopment"),
        ("MD_ProgressCode", "008", "proposed"),
        ("MD_ProgressCode", "009", "final"),
        ("MD_ProgressCode", "010", "pending"),
        ("MD_ProgressCode", "011", "retired"),
        ("MD_ProgressCode", "012", "superseded"),
        ("MD_ProgressCode", "013", "tentative"),
        ("MD_ProgressCode", "014", "valid"),
        ("MD_ProgressCode", "015", "accepted"),
        ("MD_ProgressCode", "016", "notAccepted"),
        ("MD_ProgressCode", "017", "withdrawn"),
        ("MD_ProgressCode", "018", "deprecated"),
        ("MD_RestrictionCode", "001", "copyright"),
        ("MD_RestrictionCode", "002", "patent"),
        ("MD_RestrictionCode", "003", "patentPending"),
        ("MD_RestrictionCode", "004", "trademark"),
        ("MD_RestrictionCode", "005", "license"),
        ("MD_RestrictionCode", "006", "intellectualPropertyRights"),
        ("MD_RestrictionCode", "007", "restricted"),
        ("MD_RestrictionCode", "008", "otherRestrictions"),
        ("MD_RestrictionCode", "009", "licenseUnrestricted"),
        ("MD_RestrictionCode", "010", "licenseEndUser"),
        ("MD_RestrictionCode", "011", "licenseDistributor"),
        ("MD_RestrictionCode", "012", "privacy"),
        ("MD_RestrictionCode", "013", "statutory"),
        ("MD_RestrictionCode", "014", "confidential"),
        ("MD_RestrictionCode", "015", "sensitivity/sensitiveButUnclassified"),
        ("MD_RestrictionCode", "016", "unrestricted"),
        ("MD_RestrictionCode", "017", "in-confidence"),
        ("MD_ScopeCode", "001", "attribute"),
        ("MD_ScopeCode", "002", "attributeType"),
        ("MD_ScopeCode", "003", "collectionHardware"),
        ("MD_ScopeCode", "004", "collectionSession"),
        ("MD_ScopeCode", "005", "dataset"),
        ("MD_ScopeCode", "006", "series"),
        ("MD_ScopeCode", "007", "nonGeographicDataset"),
        ("MD_ScopeCode", "008", "dimensionGroup"),
        ("MD_ScopeCode", "009", "feature"),
        ("MD_ScopeCode", "010", "featureType"),
        ("MD_ScopeCode", "011", "propertyType"),
        ("MD_ScopeCode", "012", "fieldSession"),
        ("MD_ScopeCode", "013", "software"),
        ("MD_ScopeCode", "014", "service"),
        ("MD_ScopeCode", "015", "model"),
        ("MD_ScopeCode", "016", "tile"),
        ("MD_ScopeCode", "017", "initiative"),
        ("MD_ScopeCode", "018", "stereomate"),
        ("MD_ScopeCode", "019", "sensor"),
        ("MD_ScopeCode", "020", "platformSeries"),
        ("MD_ScopeCode", "021", "sensorSeries"),
        ("MD_ScopeCode", "022", "productionSeries"),
        ("MD_ScopeCode", "023", "transferAggregate"),
        ("MD_ScopeCode", "024", "otherAggregate"),
        ("MD_ScopeCode", "025", "metadata"),
        ("MD_ScopeCode", "026", "sample"),
        ("MD_ScopeCode", "027", "document"),
        ("MD_ScopeCode", "028", "repository"),
        ("MD_ScopeCode", "029", "aggregate"),
        ("MD_ScopeCode", "030", "product"),
        ("MD_ScopeCode", "031", "collection"),
        ("MD_ScopeCode", "032", "coverage"),
        ("MD_ScopeCode", "033", "application"),
        ("MD_SpatialRepresentationTypeCode", "001", "vector"),
        ("MD_SpatialRepresentationTypeCode", "002", "grid"),
        ("MD_SpatialRepresentationTypeCode", "003", "textTable"),
        ("MD_SpatialRepresentationTypeCode", "004", "tin"),
        ("MD_SpatialRepresentationTypeCode", "005", "stereoModel"),
        ("MD_SpatialRepresentationTypeCode", "006", "video"),
        ("MD_TopicCategoryCode", "001", "farming"),
        ("MD_TopicCategoryCode", "002", "biota"),
        ("MD_TopicCategoryCode", "003", "boundaries"),
        ("MD_TopicCategoryCode", "004", "climatologyMeteorologyAtmosphere"),
        ("MD_TopicCategoryCode", "005", "economy"),
        ("MD_TopicCategoryCode", "006", "elevation"),
        ("MD_TopicCategoryCode", "007", "environment"),
        ("MD_TopicCategoryCode", "008", "geoscientificInformation"),
        ("MD_TopicCategoryCode", "009", "health"),
        ("MD_TopicCategoryCode", "010", "imageryBaseMapsEarthCover"),
        ("MD_TopicCategoryCode", "011", "intelligenceMilitary"),
        ("MD_TopicCategoryCode", "012", "inlandWaters"),
        ("MD_TopicCategoryCode", "013", "location"),
        ("MD_TopicCategoryCode", "014", "oceans"),
        ("MD_TopicCategoryCode", "015", "planningCadastre"),
        ("MD_TopicCategoryCode", "016", "society"),
        ("MD_TopicCategoryCode", "017", "structure"),
        ("MD_TopicCategoryCode", "018", "transportation"),
        ("MD_TopicCategoryCode", "019", "utilitiesCommunication"),
        ("MD_TopicCategoryCode", "020", "extraTerrestrial"),
        ("MD_TopicCategoryCode", "021", "disaster"),
        ("MD_TopologyLevelCode", "001", "geometryOnly"),
        ("MD_TopologyLevelCode", "002", "topology1D"),
        ("MD_TopologyLevelCode", "003", "planarGraph"),
        ("MD_TopologyLevelCode", "004", "fullPlanarGraph"),
        ("MD_TopologyLevelCode", "005", "surfaceGraph"),
        ("MD_TopologyLevelCode", "006", "fullSurfaceGraph"),
        ("MD_TopologyLevelCode", "007", "topology3D"),
        ("MD_TopologyLevelCode", "008", "fullTopology3D"),
        ("MD_TopologyLevelCode", "009", "abstract"),
        ("MD_CharacterSetCode", "001", "ucs2"),
        ("MD_CharacterSetCode", "002", "ucs4"),
        ("MD_CharacterSetCode", "003", "utf7"),
        ("MD_CharacterSetCode", "004", "utf8"),
        ("MD_CharacterSetCode", "005", "utf16"),
        ("MD_CharacterSetCode", "006", "8859part1"),
        ("MD_CharacterSetCode", "007", "8859part2"),
        ("MD_CharacterSetCode", "008", "8859part3"),
        ("MD_CharacterSetCode", "009", "8859part4"),
        ("MD_CharacterSetCode", "010", "8859part5"),
        ("MD_CharacterSetCode", "011", "8859part6"),
        ("MD_CharacterSetCode", "012", "8859part7"),
        ("MD_CharacterSetCode", "013", "8859part8"),
        ("MD_CharacterSetCode", "014", "8859part9"),
        ("MD_CharacterSetCode", "015", "8859part10"),
        ("MD_CharacterSetCode", "016", "8859part11"),
        ("MD_CharacterSetCode", "017", "(reserved for future use)"),
        ("MD_CharacterSetCode", "018", "8859part13"),
        ("MD_CharacterSetCode", "019", "8859part14"),
        ("MD_CharacterSetCode", "020", "8859part15"),
        ("MD_CharacterSetCode", "021", "8859part16"),
        ("MD_CharacterSetCode", "022", "jis"),
        ("MD_CharacterSetCode", "023", "shiftJIS"),
        ("MD_CharacterSetCode", "024", "eucJP"),
        ("MD_CharacterSetCode", "025", "usAscii"),
        ("MD_CharacterSetCode", "026", "ebcdic"),
        ("MD_CharacterSetCode", "027", "eucKR"),
        ("MD_CharacterSetCode", "028", "big5"),
        ("MD_CharacterSetCode", "029", "GB2312"),
        ("MD_GeometricObjectTypeCode", "001", "complex"),
        ("MD_GeometricObjectTypeCode", "002", "composite"),
        ("MD_GeometricObjectTypeCode", "003", "curve"),
        ("MD_GeometricObjectTypeCode", "004", "point"),
        ("MD_GeometricObjectTypeCode", "005", "solid"),
        ("MD_GeometricObjectTypeCode", "006", "surface"),
    ]


def _build_by_num_from_arcgis(codelist_name, by_name, arcgis_coded_values):
    """
    Build by_num (int -> display label) from ArcGIS coded values list.
    Uses by_name to resolve standard code to label; falls back to _format_code_as_label.
    """
    by_num = {}
    for name, arc_code, std_code in arcgis_coded_values:
        if name != codelist_name:
            continue
        try:
            num = int(arc_code)
        except ValueError:
            continue
        key = _normalise_code(std_code)
        label = by_name.get(key) if key else None
        if label is None:
            label = _format_code_as_label(std_code)
        by_num[num] = label
    return by_num


# Resolve ArcGIS coded values once at import (Excel if present, else inlined).
_ARCGIS_CODED_VALUES = _load_arcgis_coded_values_from_excel(
    Path(__file__).parent / "ArcGIS Metadata Details 20211130.xlsx"
) or _get_inlined_arcgis_coded_values()


def _codelist_restriction():
    # MD_RestrictionCode: limitation on access or use
    by_name = {
        "copyright": "Copyright",
        "patent": "Patent",
        "patentpending": "Patent pending",
        "trademark": "Trademark",
        "license": "Licence",
        "licence": "Licence",
        "intellectualpropertyrights": "Intellectual property rights",
        "restricted": "Restricted",
        "otherrestrictions": "Other restrictions",
        "unrestricted": "Unrestricted",
        "licenceunrestricted": "Licence unrestricted",
        "licenceenduser": "Licence end user",
        "licencedistributor": "Licence distributor",
        "private": "Private",
        "privacy": "Private",
        "statutory": "Statutory",
        "confidential": "Confidential",
        "sbu": "Sensitive but unclassified",
        "sensitivebutunclassified": "Sensitive but unclassified",
        "in-confidence": "In confidence",
    }
    by_num = _build_by_num_from_arcgis("MD_RestrictionCode", by_name, _ARCGIS_CODED_VALUES)
    return by_name, by_num


def _codelist_role():
    # CI_RoleCode: function performed by responsible party
    by_name = {
        "resourceprovider": "Resource provider",
        "custodian": "Custodian",
        "owner": "Owner",
        "sponsor": "Sponsor",
        "user": "User",
        "distributor": "Distributor",
        "originator": "Originator",
        "pointofcontact": "Point of contact",
        "principalinvestigator": "Principal investigator",
        "processor": "Processor",
        "publisher": "Publisher",
        "author": "Author",
        "coauthor": "Co-author",
        "collaborator": "Collaborator",
        "editor": "Editor",
        "mediator": "Mediator",
        "rightsholder": "Rights holder",
        "contributor": "Contributor",
        "funder": "Funder",
        "stakeholder": "Stakeholder",
    }
    by_num = _build_by_num_from_arcgis("CI_RoleCode", by_name, _ARCGIS_CODED_VALUES)
    return by_name, by_num


def _codelist_progress():
    # MD_ProgressCode: status of the dataset
    by_name = {
        "completed": "Completed",
        "historicalarchive": "Historical archive",
        "obsolete": "Obsolete",
        "ongoing": "On-going",
        "planned": "Planned",
        "required": "Required",
        "underdevelopment": "Under development",
        "final": "Final",
        "pending": "Pending",
        "retired": "Retired",
        "superseded": "Superseded",
        "tentative": "Tentative",
        "valid": "Valid",
        "accepted": "Accepted",
        "notaccepted": "Not accepted",
        "withdrawn": "Withdrawn",
        "proposed": "Proposed",
        "deprecated": "Deprecated",
    }
    by_num = _build_by_num_from_arcgis("MD_ProgressCode", by_name, _ARCGIS_CODED_VALUES)
    return by_name, by_num


def _codelist_maintenance_frequency():
    # MD_MaintenanceFrequencyCode
    by_name = {
        "continual": "Continual",
        "daily": "Daily",
        "weekly": "Weekly",
        "fortnightly": "Fortnightly",
        "monthly": "Monthly",
        "quarterly": "Quarterly",
        "biannually": "Biannually",
        "annually": "Annually",
        "asneeded": "As needed",
        "irregular": "Irregular",
        "notplanned": "Not planned",
        "unknown": "Unknown",
        "semimonthly": "Semi-monthly",
        "periodic": "Periodic",
        "biennially": "Biennially",
    }
    by_num = _build_by_num_from_arcgis(
        "MD_MaintenanceFrequencyCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_topic_category():
    # MD_TopicCategoryCode (high-level thematic classification)
    by_name = {
        "farming": "Farming",
        "biota": "Biota",
        "boundaries": "Boundaries",
        "climatologymeteorologyatmosphere": "Climatology, meteorology, atmosphere",
        "economy": "Economy",
        "elevation": "Elevation",
        "environment": "Environment",
        "geoscientificinformation": "Geoscientific information",
        "health": "Health",
        "imagerybasemapsearthcover": "Imagery, base maps, earth cover",
        "intelligencemilitary": "Intelligence, military",
        "inlandwaters": "Inland waters",
        "location": "Location",
        "oceans": "Oceans",
        "planningcadastre": "Planning, cadastre",
        "society": "Society",
        "structure": "Structure",
        "transportation": "Transportation",
        "utilitiescommunication": "Utilities, communication",
        "extraterrestrial": "Extra-terrestrial",
        "disaster": "Disaster",
    }
    by_num = _build_by_num_from_arcgis(
        "MD_TopicCategoryCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_scope():
    # MD_ScopeCode: class of information
    by_name = {
        "attribute": "Attribute",
        "attributetype": "Attribute type",
        "collectionhardware": "Collection hardware",
        "collectionsession": "Collection session",
        "dataset": "Dataset",
        "series": "Series",
        "nongeographicdataset": "Non-geographic dataset",
        "dimensiongroup": "Dimension group",
        "feature": "Feature",
        "featuretype": "Feature type",
        "propertytype": "Property type",
        "fieldsession": "Field session",
        "software": "Software",
        "service": "Service",
        "model": "Model",
        "tile": "Tile",
        "metadata": "Metadata",
        "initiative": "Initiative",
        "sample": "Sample",
        "document": "Document",
        "repository": "Repository",
        "aggregate": "Aggregate",
        "product": "Product",
        "collection": "Collection",
        "coverage": "Coverage",
        "application": "Application",
        "stereomate": "Stereomate",
        "sensor": "Sensor",
        "platformseries": "Platform series",
        "sensorseries": "Sensor series",
        "productionseries": "Production series",
        "transferaggregate": "Transfer aggregate",
        "otheraggregate": "Other aggregate",
    }
    by_num = _build_by_num_from_arcgis("MD_ScopeCode", by_name, _ARCGIS_CODED_VALUES)
    return by_name, by_num


def _codelist_character_set():
    # MD_CharacterSetCode
    by_name = {
        "ucs2": "UCS-2",
        "ucs4": "UCS-4",
        "utf7": "UTF-7",
        "utf8": "UTF-8",
        "utf16": "UTF-16",
        "8859part1": "ISO 8859-1",
        "8859part2": "ISO 8859-2",
        "8859part3": "ISO 8859-3",
        "8859part4": "ISO 8859-4",
        "8859part5": "ISO 8859-5",
        "8859part6": "ISO 8859-6",
        "8859part7": "ISO 8859-7",
        "8859part8": "ISO 8859-8",
        "8859part9": "ISO 8859-9",
        "8859part10": "ISO 8859-10",
        "8859part11": "ISO 8859-11",
        "8859part13": "ISO 8859-13",
        "8859part14": "ISO 8859-14",
        "8859part15": "ISO 8859-15",
        "8859part16": "ISO 8859-16",
        "usascii": "US ASCII",
        "ebcdic": "EBCDIC",
        "jis": "JIS",
        "shiftjis": "Shift JIS",
        "eucjp": "EUC-JP",
        "euckr": "EUC-KR",
        "big5": "Big 5",
        "gb2312": "GB 2312",
    }
    by_num = _build_by_num_from_arcgis(
        "MD_CharacterSetCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_spatial_representation():
    # MD_SpatialRepresentationTypeCode
    by_name = {
        "vector": "Vector",
        "grid": "Grid",
        "texttable": "Text, table",
        "tin": "TIN",
        "stereomodel": "Stereo model",
        "video": "Video",
    }
    by_num = _build_by_num_from_arcgis(
        "MD_SpatialRepresentationTypeCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_topology_level():
    # MD_TopologyLevelCode
    by_name = {
        "geometryonly": "Geometry only",
        "topology1d": "Topology 1D",
        "planargraph": "Planar graph",
        "fullplanargraph": "Full planar graph",
        "surfacegraph": "Surface graph",
        "fullsurfacegraph": "Full surface graph",
        "topology3d": "Topology 3D",
        "fulltopology3d": "Full topology 3D",
        "abstract": "Abstract",
    }
    by_num = _build_by_num_from_arcgis(
        "MD_TopologyLevelCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_presentation_form():
    # CI_PresentationFormCode
    by_name = {
        "documentdigital": "Document (digital)",
        "documenthardcopy": "Document (hard copy)",
        "imagedigital": "Image (digital)",
        "imagehardcopy": "Image (hard copy)",
        "mapdigital": "Map (digital)",
        "maphardcopy": "Map (hard copy)",
        "modeldigital": "Model (digital)",
        "modelhardcopy": "Model (hard copy)",
        "profiledigital": "Profile (digital)",
        "profilehardcopy": "Profile (hard copy)",
        "tabledigital": "Table (digital)",
        "tablehardcopy": "Table (hard copy)",
        "videodigital": "Video (digital)",
        "videohardcopy": "Video (hard copy)",
        "audiodigital": "Audio (digital)",
        "audiohardcopy": "Audio (hard copy)",
        "multimediadigital": "Multimedia (digital)",
        "multimediahardcopy": "Multimedia (hard copy)",
        "diagramdigital": "Diagram (digital)",
        "diagramhardcopy": "Diagram (hard copy)",
        "physicalobject": "Physical object",
    }
    by_num = _build_by_num_from_arcgis(
        "CI_PresentationFormCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_geometric_object_type():
    # MD_GeometricObjectTypeCode (geometry type of features)
    by_name = {
        "complex": "Complex",
        "composite": "Composite",
        "curve": "Curve",
        "point": "Point",
        "solid": "Solid",
        "surface": "Surface",
    }
    by_num = _build_by_num_from_arcgis(
        "MD_GeometricObjectTypeCode", by_name, _ARCGIS_CODED_VALUES
    )
    return by_name, by_num


def _codelist_content_type():
    # ArcGIS item content type (imsContentType); from Esri DTD comment.
    by_name = {
        "livedataandmaps": "Live Data and Maps",
        "downloadabledata": "Downloadable Data",
        "offlinedata": "Offline Data",
        "staticmapimages": "Static Map Images",
        "otherdocuments": "Other Documents",
        "applications": "Applications",
        "geographicservices": "Geographic Services",
        "clearinghouses": "Clearinghouses",
        "mapfiles": "Map Files",
        "geographicactivities": "Geographic Activities",
    }
    by_num = {
        1: "Live Data and Maps",
        2: "Downloadable Data",
        3: "Offline Data",
        4: "Static Map Images",
        5: "Other Documents",
        6: "Applications",
        7: "Geographic Services",
        8: "Clearinghouses",
        9: "Map Files",
        10: "Geographic Activities",
    }
    return by_name, by_num


# Codelist registry: name -> (by_name dict, by_num dict)
_CODELISTS = {
    "MD_RestrictionCode": _codelist_restriction(),
    "CI_RoleCode": _codelist_role(),
    "MD_ProgressCode": _codelist_progress(),
    "MD_MaintenanceFrequencyCode": _codelist_maintenance_frequency(),
    "MD_TopicCategoryCode": _codelist_topic_category(),
    "MD_ScopeCode": _codelist_scope(),
    "MD_CharacterSetCode": _codelist_character_set(),
    "MD_SpatialRepresentationTypeCode": _codelist_spatial_representation(),
    "MD_TopologyLevelCode": _codelist_topology_level(),
    "CI_PresentationFormCode": _codelist_presentation_form(),
    "MD_GeometricObjectTypeCode": _codelist_geometric_object_type(),
    "ArcGIS_ContentTypeCode": _codelist_content_type(),
}

# Export fields that are resolved via codelists (for the Code resolution worksheet).
# Order: (Export field name, Codelist name).
FIELD_TO_CODELIST = [
    ("Access Constraints", "MD_RestrictionCode"),
    ("Presentation Form", "CI_PresentationFormCode"),
    ("Character Set", "MD_CharacterSetCode"),
    ("Spatial Representation Type", "MD_SpatialRepresentationTypeCode"),
    ("Status", "MD_ProgressCode"),
    ("Maintenance Frequency", "MD_MaintenanceFrequencyCode"),
    ("Topic Category", "MD_TopicCategoryCode"),
    ("Contact Role", "CI_RoleCode"),
    ("Topology Level", "MD_TopologyLevelCode"),
    ("Geometry Object Type", "MD_GeometricObjectTypeCode"),
    ("Feature Geometry Code", "MD_GeometricObjectTypeCode"),
    ("Data Quality Scope Level", "MD_ScopeCode"),
    ("Metadata Maintenance Frequency", "MD_MaintenanceFrequencyCode"),
    ("Metadata Scope Code", "MD_ScopeCode"),
    ("Metadata Character Set", "MD_CharacterSetCode"),
    ("Content Type", "ArcGIS_ContentTypeCode"),
]


def get_codelist_resolution_table():
    """
    Return a list of (codelist_name, code, label) for building the Code resolution sheet.
    Code is the numeric form (e.g. 1, 5); label is the human-readable text.
    """
    rows = []
    for codelist_name in sorted(_CODELISTS.keys()):
        _by_name, by_num = _CODELISTS[codelist_name]
        for num in sorted(by_num.keys()):
            rows.append((codelist_name, str(num), by_num[num]))
    return rows


def resolve_codelist(raw_value, codelist_name):
    """
    Return a human-readable label for an ISO 19139 codelist value.

    ArcGIS often stores a numeric code (e.g. "005" for licence); standard ISO may
    use the code name (e.g. "license"). If the value is recognised, returns the
    label; otherwise returns the original value unchanged.

    Args:
        raw_value: The 'value' or 'codeListValue' from the XML (e.g. "005", "license").
        codelist_name: One of the keys in _CODELISTS (e.g. "MD_RestrictionCode").

    Returns:
        Display string (e.g. "Licence") or raw_value if not found.
    """
    if not raw_value or not isinstance(raw_value, str):
        return raw_value or ""
    raw = raw_value.strip()
    if not raw:
        return ""
    if codelist_name not in _CODELISTS:
        return raw
    by_name, by_num = _CODELISTS[codelist_name]
    # Try as name (case-insensitive, no spaces/hyphens)
    key = re.sub(r"[\s\-]", "", raw).lower()
    if key in by_name:
        return by_name[key]
    # Try as integer (strip leading zeros for lookup)
    try:
        n = int(raw)
        if n in by_num:
            return by_num[n]
    except ValueError:
        pass
    return raw


def get_field_obligation(field_name):
    """
    Return the ISO 19139/INSPIRE obligation for an exported field name.

    Used to populate the optionality row and to determine mandatory fields for
    the compliance summary. Any field not in FIELD_OBLIGATION (including
    "Other Keywords (...)" variants) is treated as optional.

    Args:
        field_name: The display name of the attribute as used in the export (e.g.
                    "Resource Title", "Abstract").

    Returns:
        One of "mandatory", "optional", or "conditional".
    """
    if field_name in FIELD_OBLIGATION:
        return FIELD_OBLIGATION[field_name]
    if field_name.startswith("Other Keywords ("):
        return "optional"
    return "optional"


def clean_text(text):
    """
    Normalise text by decoding HTML entities, stripping tags, and collapsing whitespace.

    Args:
        text: Raw string possibly containing HTML (e.g. from idAbs or useLimit).
              Can be None.

    Returns:
        Cleaned non-None string; empty string if input is None or empty after cleaning.
    """
    if text is None:
        return ""
    # Decode HTML entities
    text = html.unescape(text)
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', '', text)
    # Clean up whitespace
    text = ' '.join(text.split())
    return text.strip()


def get_text(element, default=""):
    """
    Extract and clean all text content from an XML element, including nested children.

    Handles elements that contain both direct text and child elements (e.g. div with
    spans). Result is passed through clean_text() to remove HTML and normalise space.

    Args:
        element: An xml.etree.ElementTree.Element, or None.
        default: Value to return if element is None or has no text content.

    Returns:
        Cleaned string of concatenated text; default if element is None or empty.
    """
    if element is None:
        return default
    if element.text:
        return clean_text(element.text)
    # Get all text content including nested elements
    text_parts = []
    if element.text:
        text_parts.append(element.text.strip())
    for child in element:
        if child.text:
            text_parts.append(child.text.strip())
        if child.tail:
            text_parts.append(child.tail.strip())
    return clean_text(' '.join(text_parts))


def get_attribute_value(element, attr_name, default=""):
    """
    Return the value of an XML attribute, or a default if missing or element is None.

    Args:
        element: An xml.etree.ElementTree.Element, or None.
        attr_name: Attribute name (e.g. 'value', 'codeListValue').
        default: Value to return when element is None or attribute is absent.

    Returns:
        Attribute value string, or default.
    """
    if element is None:
        return default
    return element.get(attr_name, default)


def extract_all_fields(root):
    """
    Extract all supported metadata fields from an ISO 19139 / ArcGIS metadata XML root.

    Walks Esri, dataIdInfo, mdContact, eainfo, spatRepInfo, refSysInfo, dqInfo,
    distInfo, mdMaint, mdLang, mdHrLv, spdoinfo, and related elements. Only non-empty
    values are stored. If the same field is set more than once (e.g. multiple keywords),
    values are concatenated with " | ".

    Args:
        root: The root Element of the parsed XML (e.g. from ET.parse(path).getroot()).
              Can be Esri/ArcGIS-style or standard ISO 19139.

    Returns:
        OrderedDict mapping display field names (e.g. "Resource Title") to string values.
        Order of insertion is preserved for consistent column ordering when combined
        with other files.
    """
    fields = OrderedDict()

    def add_field(field_name, value):
        """Add a non-empty field to the accumulator. If key exists, append with ' | '."""
        if value:
            # If field already exists, append with separator
            if field_name in fields:
                if fields[field_name]:
                    fields[field_name] = f"{fields[field_name]} | {value}"
                else:
                    fields[field_name] = value
            else:
                fields[field_name] = str(value)
    
    # Extract Esri metadata
    esri = root.find('.//Esri')
    if esri is not None:
        add_field("ArcGIS Format", get_text(esri.find('ArcGISFormat')))
        add_field("ArcGIS Profile", get_text(esri.find('ArcGISProfile')))
        add_field("Creation Date", get_text(esri.find('CreaDate')))
        add_field("Creation Time", get_text(esri.find('CreaTime')))
        add_field("Modification Date", get_text(esri.find('ModDate')))
        add_field("Modification Time", get_text(esri.find('ModTime')))
        
        # Data Properties
        data_props = esri.find('.//DataProperties')
        if data_props is not None:
            item_props = data_props.find('.//itemProps')
            if item_props is not None:
                add_field("Item Name", get_text(item_props.find('itemName')))
                add_field("Content Type", resolve_codelist(
                    get_text(item_props.find('imsContentType')) or "",
                    "ArcGIS_ContentTypeCode"))
                
                # Native extent
                native_ext = item_props.find('.//nativeExtBox')
                if native_ext is not None:
                    add_field("West Bounding Longitude", get_text(native_ext.find('westBL')))
                    add_field("East Bounding Longitude", get_text(native_ext.find('eastBL')))
                    add_field("South Bounding Latitude", get_text(native_ext.find('southBL')))
                    add_field("North Bounding Latitude", get_text(native_ext.find('northBL')))
                
                # Portal details
                portal = item_props.find('.//portalDetails')
                if portal is not None:
                    add_field("Thumbnail URL", get_text(portal.find('thumbnailURL')))
            
            # Coordinate reference
            coord_ref = data_props.find('.//coordRef')
            if coord_ref is not None:
                add_field("Coordinate System Type", get_text(coord_ref.find('type')))
                add_field("Geographic CS Name", get_text(coord_ref.find('geogcsn')))
                add_field("Projected CS Name", get_text(coord_ref.find('projcsn')))
                add_field("Coordinate System Units", get_text(coord_ref.find('csUnits')))
        
        # Scale range
        scale_range = esri.find('.//scaleRange')
        if scale_range is not None:
            add_field("Minimum Scale", get_text(scale_range.find('minScale')))
            add_field("Maximum Scale", get_text(scale_range.find('maxScale')))
    
    # Extract Data Identification Info
    data_id = root.find('.//dataIdInfo')
    if data_id is not None:
        # Abstract
        abstract = data_id.find('idAbs')
        if abstract is not None:
            add_field("Abstract", get_text(abstract))
        
        # Citation
        citation = data_id.find('idCitation')
        if citation is not None:
            add_field("Resource Title", get_text(citation.find('resTitle')))
            add_field("Resource Alternative Title", get_text(citation.find('resAltTitle')))
            add_field("Collection Title", get_text(citation.find('collTitle')))
            
            date_elem = citation.find('.//date/pubDate')
            if date_elem is not None:
                add_field("Publication Date", get_text(date_elem))
            
            pres_form = citation.find('.//presForm/PresFormCd')
            if pres_form is not None:
                add_field("Presentation Form", resolve_codelist(
                    get_attribute_value(pres_form, 'value') or get_attribute_value(pres_form, 'codeListValue'),
                    "CI_PresentationFormCode"))
        
        # Extent
        data_ext = data_id.find('dataExt')
        if data_ext is not None:
            add_field("Extent Description", get_text(data_ext.find('exDesc')))
            
            geo_bbox = data_ext.find('.//GeoBndBox')
            if geo_bbox is not None:
                add_field("Geographic West Bounding Longitude", get_text(geo_bbox.find('westBL')))
                add_field("Geographic East Bounding Longitude", get_text(geo_bbox.find('eastBL')))
                add_field("Geographic North Bounding Latitude", get_text(geo_bbox.find('northBL')))
                add_field("Geographic South Bounding Latitude", get_text(geo_bbox.find('southBL')))
        
        # Keywords
        keywords = data_id.findall('.//searchKeys/keyword')
        keyword_list = [get_text(kw) for kw in keywords if get_text(kw)]
        if keyword_list:
            add_field("Keywords", ', '.join(keyword_list))
        
        # Purpose
        add_field("Purpose", get_text(data_id.find('idPurp')))
        
        # Credit
        add_field("Credit", get_text(data_id.find('idCredit')))
        
        # Constraints
        use_limit = data_id.find('.//useLimit')
        if use_limit is not None:
            add_field("Use Limitation", get_text(use_limit))
        
        access_const = data_id.find('.//accessConsts/RestrictCd')
        if access_const is not None:
            add_field("Access Constraints", resolve_codelist(
                get_attribute_value(access_const, 'value') or get_attribute_value(access_const, 'codeListValue'),
                "MD_RestrictionCode"))
        
        other_const = data_id.find('.//othConsts')
        if other_const is not None:
            add_field("Other Constraints", get_text(other_const))
        
        # Language
        lang_code = data_id.find('.//dataLang/languageCode')
        if lang_code is not None:
            add_field("Data Language", get_attribute_value(lang_code, 'value'))
        
        country_code = data_id.find('.//dataLang/countryCode')
        if country_code is not None:
            add_field("Data Country Code", get_attribute_value(country_code, 'value'))
        
        # Character Set
        char_set = data_id.find('.//dataChar/CharSetCd')
        if char_set is not None:
            add_field("Character Set", resolve_codelist(
                get_attribute_value(char_set, 'value') or get_attribute_value(char_set, 'codeListValue'),
                "MD_CharacterSetCode"))
        
        # Spatial Representation Type
        spat_rep = data_id.find('.//spatRpType/SpatRepTypCd')
        if spat_rep is not None:
            add_field("Spatial Representation Type", resolve_codelist(
                get_attribute_value(spat_rep, 'value') or get_attribute_value(spat_rep, 'codeListValue'),
                "MD_SpatialRepresentationTypeCode"))
        
        # Scale
        scale = data_id.find('.//dataScale/equScale/rfDenom')
        if scale is not None:
            add_field("Scale Denominator", get_text(scale))
        
        # Environment
        envir = data_id.find('envirDesc')
        if envir is not None:
            add_field("Environment Description", get_text(envir))
        
        # Status
        status = data_id.find('.//idStatus/ProgCd')
        if status is not None:
            add_field("Status", resolve_codelist(
                get_attribute_value(status, 'value') or get_attribute_value(status, 'codeListValue'),
                "MD_ProgressCode"))
        
        # Graphic Overview
        graph_over = data_id.find('graphOver')
        if graph_over is not None:
            add_field("Graphic File Name", get_text(graph_over.find('bgFileName')))
            add_field("Graphic File Description", get_text(graph_over.find('bgFileDesc')))
            add_field("Graphic File Type", get_text(graph_over.find('bgFileType')))
        
        # Maintenance
        maint = data_id.find('.//resMaint/maintFreq/MaintFreqCd')
        if maint is not None:
            add_field("Maintenance Frequency", resolve_codelist(
                get_attribute_value(maint, 'value') or get_attribute_value(maint, 'codeListValue'),
                "MD_MaintenanceFrequencyCode"))
        
        # Topic Category
        topic_cat = data_id.find('.//tpCat/TopicCatCd')
        if topic_cat is not None:
            add_field("Topic Category", resolve_codelist(
                get_attribute_value(topic_cat, 'value') or get_attribute_value(topic_cat, 'codeListValue'),
                "MD_TopicCategoryCode"))
        
        # Other Keywords
        other_keys = data_id.findall('.//otherKeys')
        for other_key in other_keys:
            thesa_name = get_text(other_key.find('.//thesaName/resTitle'))
            keywords = [get_text(kw) for kw in other_key.findall('keyword') if get_text(kw)]
            if keywords:
                key_name = f"Other Keywords ({thesa_name})" if thesa_name else "Other Keywords"
                add_field(key_name, ', '.join(keywords))
    
    # Extract Contact Information
    contact = root.find('.//mdContact')
    if contact is not None:
        add_field("Contact Individual Name", get_text(contact.find('rpIndName')))
        add_field("Contact Organisation Name", get_text(contact.find('rpOrgName')))
        add_field("Contact Position Name", get_text(contact.find('rpPosName')))
        
        cnt_info = contact.find('.//rpCntInfo')
        if cnt_info is not None:
            # Address
            address = cnt_info.find('.//cntAddress')
            if address is not None:
                add_field("Contact Email Address", get_text(address.find('eMailAdd')))
                add_field("Contact Delivery Point", get_text(address.find('delPoint')))
                add_field("Contact City", get_text(address.find('city')))
                add_field("Contact Administrative Area", get_text(address.find('adminArea')))
                add_field("Contact Postal Code", get_text(address.find('postCode')))
                add_field("Contact Country", get_text(address.find('country')))
            
            # Phone
            phone = cnt_info.find('.//cntPhone/voiceNum')
            if phone is not None:
                add_field("Contact Phone Number", get_text(phone))
            
            # Online Resource
            online = cnt_info.find('.//cntOnlineRes/linkage')
            if online is not None:
                add_field("Contact Online Resource", get_text(online))
            
            # Hours
            hours = cnt_info.find('.//cntHours')
            if hours is not None:
                add_field("Contact Hours", get_text(hours))
            
            # Instructions
            instr = cnt_info.find('.//cntInstr')
            if instr is not None:
                add_field("Contact Instructions", get_text(instr))
        
        role = contact.find('.//role/RoleCd')
        if role is not None:
            add_field("Contact Role", resolve_codelist(
                get_attribute_value(role, 'value') or get_attribute_value(role, 'codeListValue'),
                "CI_RoleCode"))
    
    # Extract Attribute Definitions (eainfo)
    eainfo = root.find('.//eainfo/detailed')
    if eainfo is not None:
        enttyp = eainfo.find('enttyp')
        if enttyp is not None:
            add_field("Entity Type Label", get_text(enttyp.find('enttypl')))
            add_field("Entity Type Type", get_text(enttyp.find('enttypt')))
            add_field("Entity Type Count", get_text(enttyp.find('enttypc')))
        
        # Process all attributes - store as a summary
        attributes = eainfo.findall('.//attr')
        attr_summaries = []
        for attr in attributes:
            attr_label = get_text(attr.find('attrlabl'))
            if attr_label:
                attr_summaries.append(attr_label)
        if attr_summaries:
            add_field("Attribute Names", ', '.join(attr_summaries))
    
    # Extract Spatial Representation Info
    spat_rep_info = root.find('.//spatRepInfo')
    if spat_rep_info is not None:
        top_level = spat_rep_info.find('.//topLvl/TopoLevCd')
        if top_level is not None:
            add_field("Topology Level", resolve_codelist(
                get_attribute_value(top_level, 'value') or get_attribute_value(top_level, 'codeListValue'),
                "MD_TopologyLevelCode"))
        
        geo_objs = spat_rep_info.find('.//geometObjs')
        if geo_objs is not None:
            geo_type = geo_objs.find('.//geoObjTyp/GeoObjTypCd')
            if geo_type is not None:
                add_field("Geometry Object Type", resolve_codelist(
                    get_attribute_value(geo_type, 'value') or get_attribute_value(geo_type, 'codeListValue'),
                    "MD_GeometricObjectTypeCode"))
            
            geo_count = geo_objs.find('.//geoObjCnt')
            if geo_count is not None:
                add_field("Geometry Object Count", get_text(geo_count))
    
    # Extract Reference System Info
    ref_sys = root.find('.//refSysInfo/RefSystem/refSysID')
    if ref_sys is not None:
        ident_code = ref_sys.find('identCode')
        if ident_code is not None:
            add_field("Reference System Code", get_attribute_value(ident_code, 'code'))
        
        code_space = ref_sys.find('idCodeSpace')
        if code_space is not None:
            add_field("Reference System Code Space", get_text(code_space))
        
        version = ref_sys.find('idVersion')
        if version is not None:
            add_field("Reference System Version", get_text(version))
    
    # Extract Data Quality Info
    dq_info = root.find('.//dqInfo')
    if dq_info is not None:
        scope = dq_info.find('.//scpLvl/ScopeCd')
        if scope is not None:
            add_field("Data Quality Scope Level", resolve_codelist(
                get_attribute_value(scope, 'value') or get_attribute_value(scope, 'codeListValue'),
                "MD_ScopeCode"))
        
        lineage = dq_info.find('.//dataLineage/statement')
        if lineage is not None:
            add_field("Lineage Statement", get_text(lineage))
        
        report = dq_info.find('.//report')
        if report is not None:
            report_type = report.get('type', '')
            add_field("Quality Report Type", report_type)
            
            con_spec = report.find('.//conSpec/resTitle')
            if con_spec is not None:
                add_field("Conformance Specification Title", get_text(con_spec))
            
            con_pass = report.find('.//conPass')
            if con_pass is not None:
                add_field("Conformance Pass", get_text(con_pass))
    
    # Extract Distribution Info
    dist_info = root.find('.//distInfo')
    if dist_info is not None:
        online_src = dist_info.find('.//onLineSrc')
        if online_src is not None:
            linkage = online_src.find('linkage')
            if linkage is not None:
                add_field("Distribution Online Resource Linkage", get_text(linkage))
            
            protocol = online_src.find('protocol')
            if protocol is not None:
                add_field("Distribution Protocol", get_text(protocol))
            
            or_name = online_src.find('orName')
            if or_name is not None:
                add_field("Distribution Online Resource Name", get_text(or_name))
            
            or_desc = online_src.find('orDesc')
            if or_desc is not None:
                add_field("Distribution Online Resource Description", get_text(or_desc))
    
    # Extract Maintenance Info
    md_maint = root.find('.//mdMaint')
    if md_maint is not None:
        maint_freq = md_maint.find('.//maintFreq/MaintFreqCd')
        if maint_freq is not None:
            add_field("Metadata Maintenance Frequency", resolve_codelist(
                get_attribute_value(maint_freq, 'value') or get_attribute_value(maint_freq, 'codeListValue'),
                "MD_MaintenanceFrequencyCode"))
    
    # Extract Metadata Language
    md_lang = root.find('.//mdLang')
    if md_lang is not None:
        lang_code = md_lang.find('languageCode')
        if lang_code is not None:
            add_field("Metadata Language Code", get_attribute_value(lang_code, 'value'))
        
        country_code = md_lang.find('countryCode')
        if country_code is not None:
            add_field("Metadata Country Code", get_attribute_value(country_code, 'value'))
    
    # Extract Metadata Hierarchy Level
    md_hr_lv = root.find('.//mdHrLv')
    if md_hr_lv is not None:
        scope_cd = md_hr_lv.find('ScopeCd')
        if scope_cd is not None:
            add_field("Metadata Scope Code", resolve_codelist(
                get_attribute_value(scope_cd, 'value') or get_attribute_value(scope_cd, 'codeListValue'),
                "MD_ScopeCode"))
    
    hr_lv_name = root.find('.//mdHrLvName')
    if hr_lv_name is not None:
        add_field("Metadata Hierarchy Level Name", get_text(hr_lv_name))
    
    # Extract Spatial Domain Info
    spdo_info = root.find('.//spdoinfo')
    if spdo_info is not None:
        esri_term = spdo_info.find('.//esriterm')
        if esri_term is not None:
            name = esri_term.get('Name', '')
            add_field("Feature Name", name)
            
            feat_type = esri_term.find('efeatyp')
            if feat_type is not None:
                add_field("Feature Type", get_text(feat_type))
            
            feat_geom = esri_term.find('efeageom')
            if feat_geom is not None:
                add_field("Feature Geometry Code", resolve_codelist(
                    get_attribute_value(feat_geom, 'code') or "",
                    "MD_GeometricObjectTypeCode"))
    
    # Extract Metadata Standard
    md_std_name = root.find('.//mdStanName')
    if md_std_name is not None:
        add_field("Metadata Standard Name", get_text(md_std_name))
    
    md_std_ver = root.find('.//mdStanVer')
    if md_std_ver is not None:
        add_field("Metadata Standard Version", get_text(md_std_ver))
    
    # Extract File ID
    md_file_id = root.find('.//mdFileID')
    if md_file_id is not None:
        add_field("Metadata File ID", get_text(md_file_id))
    
    md_char = root.find('.//mdChar')
    if md_char is not None:
        char_set = md_char.find('CharSetCd')
        if char_set is not None:
            add_field("Metadata Character Set", resolve_codelist(
                get_attribute_value(char_set, 'value') or get_attribute_value(char_set, 'codeListValue'),
                "MD_CharacterSetCode"))
    
    md_date = root.find('.//mdDateSt')
    if md_date is not None:
        add_field("Metadata Date Stamp", get_text(md_date))
    
    return fields


def process_all_xml_files(xml_folder):
    """
    Discover and process every .xml file in the given folder.

    Each file is parsed and passed to extract_all_fields(). Failures for a single
    file are logged but do not stop the run. The union of all attribute names
    across successful files is returned as the canonical column set.

    Args:
        xml_folder: Path to the directory containing .xml metadata files (str or Path).

    Returns:
        Tuple (all_data, sorted_field_names), or (None, None) if the folder does not
        exist or contains no .xml files.
        - all_data: Dict mapping filename (str) to OrderedDict of field name -> value.
        - sorted_field_names: Sorted list of all unique attribute names found.
    """
    xml_folder = Path(xml_folder)
    xml_files = sorted(xml_folder.glob("*.xml"))
    
    if not xml_files:
        print(f"No XML files found in {xml_folder}")
        return None, None
    
    print(f"Found {len(xml_files)} XML files to process")
    
    all_data = {}  # filename -> fields dictionary
    all_field_names = set()  # Collect all unique field names
    
    for xml_file in xml_files:
        filename = xml_file.name
        print(f"Processing: {filename}")
        
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            fields = extract_all_fields(root)
            
            all_data[filename] = fields
            all_field_names.update(fields.keys())
            
        except ET.ParseError as e:
            print(f"  Error parsing {filename}: {e}")
        except Exception as e:
            print(f"  Error processing {filename}: {e}")
    
    # Sort field names for consistent column order
    sorted_field_names = sorted(all_field_names)
    
    return all_data, sorted_field_names


def compute_compliance(all_data, field_names):
    """
    Assess ISO 19139 compliance for each file based on presence of mandatory fields.

    A file is considered compliant only if every field marked as "mandatory" in
    FIELD_OBLIGATION is present in that file's extracted data and has a non-empty
    value. Conditional and optional fields do not affect the compliant flag.

    Args:
        all_data: Dict mapping filename to dict of field name -> value (as from
                  process_all_xml_files).
        field_names: List of all attribute names (column set).

    Returns:
        List of dicts, one per file (sorted by filename), each with keys:
        - "Filename": str
        - "Compliant": "Yes" or "No"
        - "Missing mandatory": comma-separated list of missing mandatory field names
        - "Missing count": int
    """
    mandatory_fields = [fn for fn in field_names if get_field_obligation(fn) == "mandatory"]
    results = []
    for filename in sorted(all_data.keys()):
        fields = all_data[filename]
        missing = [fn for fn in mandatory_fields if not (fields.get(fn) or "").strip()]
        compliant = len(missing) == 0
        results.append({
            "Filename": filename,
            "Compliant": "Yes" if compliant else "No",
            "Missing mandatory": ", ".join(missing) if missing else "",
            "Missing count": len(missing),
        })
    return results


def create_excel_matrix(all_data, field_names, output_file):
    """
    Build the Excel workbook and write it to disk.

    Creates three sheets:
    1. "Compliance Summary" (first): one row per file with Filename, ISO 19139
       compliant (Yes/No), Missing mandatory fields, Missing count.
    2. "Metadata Export": Row 1 = headers (Filename + attribute names), Row 2 =
       optionality (mandatory/optional/conditional per column), Row 3+ = one row per
       file with filename and attribute values. Freezes header and optionality rows
       and the filename column. Applies styling and text wrapping.
    3. "Code resolution": Lists which export fields use codelist resolution and
       summarises how numeric/code values are mapped to human-readable labels.

    Args:
        all_data: Dict mapping filename to dict of field name -> value.
        field_names: List of attribute names defining column order.
        output_file: Path (str or Path) for the output .xlsx file.

    Returns:
        None. Writes the file and prints a short summary to stdout.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata Export"

    # Row 1: Header row – Filename + all field names
    headers = ['Filename'] + field_names
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 2: Optionality row – blank for Filename, then mandatory/optional/conditional per column
    ws.cell(row=2, column=1).value = ""  # Filename column
    obligation_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    obligation_font = Font(italic=True)
    for col_num, field_name in enumerate(field_names, 2):
        cell = ws.cell(row=2, column=col_num)
        cell.value = get_field_obligation(field_name)
        cell.fill = obligation_fill
        cell.font = obligation_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows (starting at row 3)
    for row_idx, filename in enumerate(sorted(all_data.keys()), start=3):
        fields = all_data[filename]
        ws.cell(row=row_idx, column=1).value = filename
        for col_num, field_name in enumerate(field_names, 2):
            val = fields.get(field_name, '')
            ws.cell(row=row_idx, column=col_num).value = val

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(str(header))
        column_letter = get_column_letter(col_num)
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_num)
            if cell.value and len(str(cell.value)) > max_length:
                max_length = min(len(str(cell.value)), 100)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 100)

    # Freeze header and optionality row plus filename column
    ws.freeze_panes = 'B3'

    # Text wrapping for data cells; shade empty cells light gray
    empty_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if not (cell.value is not None and str(cell.value).strip()):
                cell.fill = empty_fill

    # Compliance Summary sheet
    compliance = compute_compliance(all_data, field_names)
    ws_summary = wb.create_sheet("Compliance Summary", 0)
    summary_headers = ["Filename", "ISO 19139 compliant", "Missing mandatory fields", "Missing count"]
    for col_num, h in enumerate(summary_headers, 1):
        c = ws_summary.cell(row=1, column=col_num)
        c.value = h
        c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row_idx, rec in enumerate(compliance, start=2):
        ws_summary.cell(row=row_idx, column=1).value = rec["Filename"]
        ws_summary.cell(row=row_idx, column=2).value = rec["Compliant"]
        ws_summary.cell(row=row_idx, column=3).value = rec["Missing mandatory"]
        ws_summary.cell(row=row_idx, column=4).value = rec["Missing count"]
    for col_num in range(1, 5):
        ws_summary.column_dimensions[get_column_letter(col_num)].width = 24
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Code resolution worksheet: fields that use codelists and how codes map to text
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_codes = wb.create_sheet("Code resolution", 2)
    row_num = 1
    ws_codes.cell(row=row_num, column=1, value="Fields using code resolution - where code numbers are replaced in the report with full text for ease of reading")
    ws_codes.cell(row=row_num, column=1).font = Font(bold=True, size=12)
    row_num += 1
    ws_codes.cell(row=row_num, column=1, value="Export field name")
    ws_codes.cell(row=row_num, column=2, value="Codelist")
    for c in range(1, 3):
        ws_codes.cell(row=row_num, column=c).fill = header_fill
        ws_codes.cell(row=row_num, column=c).font = header_font
        ws_codes.cell(row=row_num, column=c).alignment = header_align
    row_num += 1
    for field_name, codelist_name in FIELD_TO_CODELIST:
        ws_codes.cell(row=row_num, column=1, value=field_name)
        ws_codes.cell(row=row_num, column=2, value=codelist_name)
        row_num += 1
    row_num += 1
    ws_codes.cell(row=row_num, column=1, value="How codes are resolved to text")
    ws_codes.cell(row=row_num, column=1).font = Font(bold=True, size=12)
    row_num += 1
    ws_codes.cell(row=row_num, column=1, value="Numeric codes (e.g. 005 in XML) and code names (e.g. license) are mapped to the labels below. ArcGIS often uses 3-digit numeric values.")
    ws_codes.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True)
    row_num += 2
    ws_codes.cell(row=row_num, column=1, value="Codelist")
    ws_codes.cell(row=row_num, column=2, value="Code (numeric or name)")
    ws_codes.cell(row=row_num, column=3, value="Resolved label")
    for c in range(1, 4):
        ws_codes.cell(row=row_num, column=c).fill = header_fill
        ws_codes.cell(row=row_num, column=c).font = header_font
        ws_codes.cell(row=row_num, column=c).alignment = header_align
    row_num += 1
    for codelist_name, code, label in get_codelist_resolution_table():
        ws_codes.cell(row=row_num, column=1, value=codelist_name)
        ws_codes.cell(row=row_num, column=2, value=code)
        ws_codes.cell(row=row_num, column=3, value=label)
        row_num += 1
    ws_codes.column_dimensions["A"].width = 32
    ws_codes.column_dimensions["B"].width = 28
    ws_codes.column_dimensions["C"].width = 36

    wb.save(output_file)
    print(f"\nExcel file created successfully: {output_file}")
    print(f"Total files processed: {len(all_data)}")
    print(f"Total unique attributes: {len(field_names)}")
    compliant_count = sum(1 for r in compliance if r["Compliant"] == "Yes")
    print(f"ISO 19139 compliance: {compliant_count} compliant, {len(compliance) - compliant_count} with missing mandatory fields")


def parse_args():
    """
    Parse command-line arguments for input folder and derive output filename.

    Returns:
        argparse.Namespace with 'input_folder' (Path) and 'output_file' (str).
        output_file is of the form metadata_export_<foldername>.xlsx so the
        Excel file name reflects the source folder.
    """
    parser = argparse.ArgumentParser(
        description="Extract metadata from ISO 19139 / ArcGIS XML files into an Excel workbook."
    )
    parser.add_argument(
        "input_folder",
        nargs="?",
        default="xml",
        help="Folder containing .xml metadata files (default: xml)",
    )
    args = parser.parse_args()
    xml_folder = Path(args.input_folder)
    folder_name = xml_folder.name
    reports_dir = Path("reports")
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_file = reports_dir / f"metadata_export_{folder_name}.xlsx"
    args.input_folder = xml_folder
    args.output_file = output_file
    return args


def main():
    """
    Entry point: run batch extraction and write an Excel workbook.

    The input folder is a script parameter (default: 'xml'). The output file is
    named metadata_export_<foldername>.xlsx so it reflects the source folder.
    Exits with an error if the folder is missing. On success, prints progress
    per file and final counts (files processed, unique attributes, compliant
    vs non-compliant).
    """
    args = parse_args()
    xml_folder = args.input_folder
    output_file = args.output_file

    if not xml_folder.exists():
        print(f"Error: XML folder not found at {xml_folder}")
        return

    print(f"Processing XML files from: {xml_folder}")
    print("-" * 60)

    try:
        all_data, field_names = process_all_xml_files(xml_folder)

        if all_data is None:
            return

        print("-" * 60)
        print(f"Creating Excel file: {output_file}")
        create_excel_matrix(all_data, field_names, output_file)

        print("\nExtraction complete!")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
