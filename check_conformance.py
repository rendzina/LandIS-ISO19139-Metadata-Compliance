#!/usr/bin/env python3
"""
Strict namespace-aware ISO 19139 conformance checker.

Processes all .xml files in a given folder and produces an Excel report showing
how each file satisfies obligatory (mandatory), optional, and conditional
elements per ISO 19139 / INSPIRE (Regulation 1205/2008). Uses the official
gmd/gco namespaces so that only properly namespaced ISO 19139 content is
recognised.

Usage:
    python check_conformance.py [folder]
    Default folder: xml. Output: conformance_report_<foldername>.xlsx
"""

import argparse
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ISO 19139 namespace URIs (ISO/TC 211)
GMD = "http://www.isotc211.org/2005/gmd"
GCO = "http://www.isotc211.org/2005/gco"


def _tag(ns, local):
    """Return Clark-notation tag for ElementTree find()."""
    return "{%s}%s" % (ns, local)


def _find(root, path):
    """
    Find an element by a path of (namespace_uri, local_name) pairs.
    Returns the element or None if any step fails.
    """
    for ns, local in path:
        if root is None:
            return None
        root = root.find(_tag(ns, local))
    return root


def _find_any(root, paths):
    """
    Try each path (list of (ns, local) steps); return first non-None element.
    """
    for path in paths:
        el = _find(root, path)
        if el is not None:
            return el
    return None


def _get_text(element):
    """
    Collect all text content from an element and its descendants (e.g. from
    gco:CharacterString, gco:Date, or code list elements). Returns stripped
    string; empty string if element is None or has no text.
    """
    if element is None:
        return ""
    parts = []
    if element.text:
        parts.append(element.text)
    for child in element:
        parts.append(_get_text(child))
        if child.tail:
            parts.append(child.tail)
    return " ".join(parts).strip()


def _has_value(element):
    """Return True if element exists and has non-empty text content."""
    return bool(_get_text(element)) if element is not None else False


# Conformance checks: (display_name, obligation, finder_callable).
# Finder receives root and returns one of: "Present", "Empty", "Absent".
# We build a list of (name, obligation, finder) where finder(root) -> str.

def _check_single(paths, root):
    """One or more paths (each path is list of (ns, local)); if any leads to element with content, Present; else Empty or Absent."""
    el = _find_any(root, paths)
    if el is None:
        return "Absent"
    return "Present" if _has_value(el) else "Empty"


def _check_bbox(root):
    """All four geographic bounding box elements must exist and have content."""
    base = [
        (GMD, "identificationInfo"),
        (GMD, "MD_DataIdentification"),
        (GMD, "extent"),
        (GMD, "EX_Extent"),
        (GMD, "geographicElement"),
        (GMD, "EX_GeographicBoundingBox"),
    ]
    for comp in ["westBoundLongitude", "eastBoundLongitude", "southBoundLatitude", "northBoundLatitude"]:
        el = _find(root, base + [(GMD, comp)])
        if el is None or not _has_value(el):
            return "Absent" if el is None else "Empty"
    return "Present"


def _check_keywords(root):
    """At least one keyword with content under descriptiveKeywords."""
    id_info = _find(root, [(GMD, "identificationInfo"), (GMD, "MD_DataIdentification")])
    if id_info is None:
        return "Absent"
    # Can have multiple descriptiveKeywords; each can have multiple keyword
    for kw_cont in id_info.findall(".//%s" % _tag(GMD, "descriptiveKeywords")):
        for kw in kw_cont.findall(".//%s" % _tag(GMD, "keyword")):
            if _has_value(kw):
                return "Present"
    return "Empty"


def _check_use_limitation(root):
    """useLimitation can appear in MD_Constraints or MD_LegalConstraints."""
    id_info = _find(root, [(GMD, "identificationInfo"), (GMD, "MD_DataIdentification")])
    if id_info is None:
        return "Absent"
    for use_lim in id_info.findall(".//%s" % _tag(GMD, "useLimitation")):
        if _has_value(use_lim):
            return "Present"
    return "Empty" if id_info.find(".//%s" % _tag(GMD, "useLimitation")) is not None else "Absent"


def _check_access_constraints(root):
    """accessConstraints in MD_LegalConstraints or RestrictCd in MD_Constraints."""
    id_info = _find(root, [(GMD, "identificationInfo"), (GMD, "MD_DataIdentification")])
    if id_info is None:
        return "Absent"
    # MD_LegalConstraints/accessConstraints
    ac = id_info.find(".//%s/%s" % (_tag(GMD, "MD_LegalConstraints"), _tag(GMD, "accessConstraints")))
    if ac is not None and _has_value(ac):
        return "Present"
    # Or RestrictCd under accessConsts (some profiles)
    restrict = id_info.find(".//%s" % _tag(GMD, "accessConstraints"))
    if restrict is not None and _has_value(restrict):
        return "Present"
    return "Empty" if (ac is not None or restrict is not None) else "Absent"


def _check_other_constraints(root):
    id_info = _find(root, [(GMD, "identificationInfo"), (GMD, "MD_DataIdentification")])
    if id_info is None:
        return "Absent"
    other = id_info.find(".//%s" % _tag(GMD, "otherConstraints"))
    if other is None:
        return "Absent"
    return "Present" if _has_value(other) else "Empty"


def _check_distribution_linkage(root):
    """Distribution: onLine/CI_OnlineResource/linkage (URL or CharacterString)."""
    dist = _find(root, [(GMD, "distributionInfo"), (GMD, "MD_Distribution")])
    if dist is None:
        return "Absent"
    for on_line in dist.findall(".//%s" % _tag(GMD, "onLine")):
        res = on_line.find(_tag(GMD, "CI_OnlineResource"))
        if res is None:
            continue
        linkage = res.find(_tag(GMD, "linkage"))
        if linkage is not None and _has_value(linkage):
            return "Present"
    return "Empty" if dist.find(".//%s" % _tag(GMD, "linkage")) is not None else "Absent"


def _check_conformance_spec_and_pass(root):
    """Conformance: specification title and pass in DQ_ConformanceResult."""
    dq = _find(root, [(GMD, "dataQualityInfo"), (GMD, "DQ_DataQuality")])
    if dq is None:
        return "Absent"
    result = dq.find(".//%s" % _tag(GMD, "DQ_ConformanceResult"))
    if result is None:
        return "Absent"
    spec = result.find(".//%s/%s" % (_tag(GMD, "specification"), _tag(GMD, "CI_Citation")))
    if spec is not None:
        title = spec.find(_tag(GMD, "title"))
        if title is not None and _has_value(title):
            pass_el = result.find(".//%s" % _tag(GCO, "pass"))
            if pass_el is not None and _has_value(pass_el):
                return "Present"
    return "Empty"


# Build list of (display_name, obligation, finder)
def _conformance_checks():
    g = GMD
    c = GCO
    # (name, obligation, finder)
    checks = [
        # Mandatory – identification
        ("Resource Title", "mandatory", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "citation"), (g, "CI_Citation"), (g, "title")]], r)),
        ("Abstract", "mandatory", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "abstract")]], r)),
        ("Topic Category", "mandatory", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "topicCategory"), (g, "MD_TopicCategoryCode")]], r)),
        ("Keywords", "mandatory", _check_keywords),
        ("Geographic bounding box", "mandatory", _check_bbox),
        ("Data Language", "mandatory", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "language")]], r)),
        ("Scale Denominator", "mandatory", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "spatialResolution"), (g, "MD_Resolution"), (g, "equivalentScale"), (g, "MD_RepresentativeFraction"), (g, "denominator")]], r)),
        # Mandatory – contact
        ("Contact Organisation Name", "mandatory", lambda r: _check_single([[(g, "contact"), (g, "CI_ResponsibleParty"), (g, "organisationName")]], r)),
        ("Contact Email Address", "mandatory", lambda r: _check_single([[(g, "contact"), (g, "CI_ResponsibleParty"), (g, "contactInfo"), (g, "CI_Contact"), (g, "address"), (g, "CI_Address"), (g, "electronicMailAddress")]], r)),
        ("Contact Role", "mandatory", lambda r: _check_single([[(g, "contact"), (g, "CI_ResponsibleParty"), (g, "role")]], r)),
        # Mandatory – distribution
        ("Distribution Online Resource Linkage", "mandatory", _check_distribution_linkage),
        # Mandatory – data quality
        ("Lineage Statement", "mandatory", lambda r: _check_single([[(g, "dataQualityInfo"), (g, "DQ_DataQuality"), (g, "lineage"), (g, "LI_Lineage"), (g, "statement")]], r)),
        ("Data Quality Scope Level", "mandatory", lambda r: _check_single([[(g, "dataQualityInfo"), (g, "DQ_DataQuality"), (g, "scope"), (g, "DQ_Scope"), (g, "level")]], r)),
        ("Conformance Specification Title", "mandatory", lambda r: _check_single([[(g, "dataQualityInfo"), (g, "DQ_DataQuality"), (g, "report"), (g, "DQ_DomainConsistency"), (g, "result"), (g, "DQ_ConformanceResult"), (g, "specification"), (g, "CI_Citation"), (g, "title")]], r)),
        ("Conformance Pass", "mandatory", lambda r: _check_single([[(g, "dataQualityInfo"), (g, "DQ_DataQuality"), (g, "report"), (g, "DQ_DomainConsistency"), (g, "result"), (g, "DQ_ConformanceResult"), (g, "pass")]], r)),
        # Mandatory – metadata section
        ("Metadata Language Code", "mandatory", lambda r: _check_single([[(g, "language")]], r)),
        ("Metadata Date Stamp", "mandatory", lambda r: _check_single([[(g, "dateStamp")]], r)),
        ("Metadata Scope Code", "mandatory", lambda r: _check_single([[(g, "hierarchyLevel")]], r)),
        # Mandatory – constraints
        ("Access Constraints", "mandatory", _check_access_constraints),
        ("Other Constraints", "mandatory", _check_other_constraints),
        ("Use Limitation", "mandatory", _check_use_limitation),
        # Conditional
        ("Publication Date", "conditional", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "citation"), (g, "CI_Citation"), (g, "date"), (g, "CI_Date"), (g, "date")]], r)),
        ("Reference System Code", "conditional", lambda r: _check_single([[(g, "referenceSystemInfo"), (g, "MD_ReferenceSystem"), (g, "referenceSystemIdentifier"), (g, "RS_Identifier"), (g, "code")]], r)),
        ("Reference System Code Space", "conditional", lambda r: _check_single([[(g, "referenceSystemInfo"), (g, "MD_ReferenceSystem"), (g, "referenceSystemIdentifier"), (g, "RS_Identifier"), (g, "codeSpace")]], r)),
        # Optional (sample)
        ("File Identifier", "optional", lambda r: _check_single([[(g, "fileIdentifier")]], r)),
        ("Metadata Standard Name", "optional", lambda r: _check_single([[(g, "metadataStandardName")]], r)),
        ("Metadata Standard Version", "optional", lambda r: _check_single([[(g, "metadataStandardVersion")]], r)),
        ("Purpose", "optional", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "purpose")]], r)),
        ("Credit", "optional", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "credit")]], r)),
        ("Status", "optional", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "status")]], r)),
        ("Maintenance Frequency", "optional", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "resourceMaintenance"), (g, "MD_MaintenanceInformation"), (g, "maintenanceAndUpdateFrequency")]], r)),
        ("Graphic Overview", "optional", lambda r: _check_single([[(g, "identificationInfo"), (g, "MD_DataIdentification"), (g, "graphicOverview"), (g, "MD_BrowseGraphic"), (g, "fileName")]], r)),
    ]
    return checks


def check_one_file(xml_path, checks):
    """
    Run all conformance checks on a single XML file using namespace-aware parsing.

    Returns:
        dict mapping check display_name to "Present" | "Empty" | "Absent".
        Returns None if the file is not valid ISO 19139 (e.g. root is not MD_Metadata in gmd namespace).
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError:
        return None

    # Require root to be MD_Metadata in gmd namespace
    if root.tag != _tag(GMD, "MD_Metadata"):
        return None

    results = {}
    for name, _obligation, finder in checks:
        try:
            results[name] = finder(root)
        except Exception:
            results[name] = "Absent"
    return results


def process_folder(folder_path, checks):
    """
    Process all .xml files in folder_path. Returns (results_dict, errors).
    results_dict: filename -> check_name -> "Present"|"Empty"|"Absent".
    errors: list of (filename, error_message).
    """
    folder_path = Path(folder_path)
    if not folder_path.exists():
        return None, ["Folder not found: %s" % folder_path]

    xml_files = sorted(folder_path.glob("*.xml"))
    if not xml_files:
        return None, ["No XML files found in %s" % folder_path]

    results = {}
    errors = []
    for xml_file in xml_files:
        name = xml_file.name
        row = check_one_file(xml_file, checks)
        if row is None:
            errors.append((name, "Not ISO 19139 namespaced (root is not gmd:MD_Metadata)"))
            continue
        results[name] = row

    return results, errors


def compute_summary(results, checks):
    """Per-file summary: conformant (all mandatory Present), missing mandatory list, counts."""
    mandatory = [name for name, obl, _ in checks if obl == "mandatory"]
    summary = []
    for filename in sorted(results.keys()):
        row = results[filename]
        missing = [n for n in mandatory if row.get(n) != "Present"]
        conformant = len(missing) == 0
        present_mandatory = sum(1 for n in mandatory if row.get(n) == "Present")
        present_conditional = sum(1 for name, obl, _ in checks if obl == "conditional" and row.get(name) == "Present")
        present_optional = sum(1 for name, obl, _ in checks if obl == "optional" and row.get(name) == "Present")
        summary.append({
            "Filename": filename,
            "Conformant": "Yes" if conformant else "No",
            "Missing mandatory": ", ".join(missing) if missing else "",
            "Missing count": len(missing),
            "Present mandatory": present_mandatory,
            "Present conditional": present_conditional,
            "Present optional": present_optional,
        })
    return summary


def write_excel_report(results, checks, summary, errors, output_path):
    """Write Excel workbook: Compliance Summary sheet and Conformance Detail sheet."""
    wb = Workbook()
    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Compliance Summary"
    headers = [
        "Filename", "ISO 19139 conformant", "Missing mandatory fields",
        "Missing count", "Present mandatory", "Present conditional", "Present optional",
    ]
    for col, h in enumerate(headers, 1):
        ws_sum.cell(row=1, column=col, value=h)
    for row_idx, rec in enumerate(summary, 2):
        ws_sum.cell(row=row_idx, column=1, value=rec["Filename"])
        ws_sum.cell(row=row_idx, column=2, value=rec["Conformant"])
        ws_sum.cell(row=row_idx, column=3, value=rec["Missing mandatory"])
        ws_sum.cell(row=row_idx, column=4, value=rec["Missing count"])
        ws_sum.cell(row=row_idx, column=5, value=rec["Present mandatory"])
        ws_sum.cell(row=row_idx, column=6, value=rec["Present conditional"])
        ws_sum.cell(row=row_idx, column=7, value=rec["Present optional"])
    for col in range(1, len(headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Detail (one column per check, one row per file)
    ws_det = wb.create_sheet("Conformance Detail", 1)
    # Row 1: check name
    # Row 2: obligation (mandatory/conditional/optional)
    # Row 3+: one row per file
    ws_det.cell(row=1, column=1, value="Filename")
    ws_det.cell(row=2, column=1, value="(obligation)")
    for col_idx, (name, obligation, _) in enumerate(checks, 2):
        ws_det.cell(row=1, column=col_idx, value=name)
        ws_det.cell(row=2, column=col_idx, value=obligation)
    for row_idx, filename in enumerate(sorted(results.keys()), start=3):
        ws_det.cell(row=row_idx, column=1, value=filename)
        row = results[filename]
        for col_idx, (name, _, _) in enumerate(checks, 2):
            ws_det.cell(row=row_idx, column=col_idx, value=row.get(name, "Absent"))
    ws_det.freeze_panes = "B3"
    for col in range(1, len(checks) + 2):
        ws_det.column_dimensions[get_column_letter(col)].width = 12

    # Optional: Errors sheet if any files were skipped
    if errors:
        ws_err = wb.create_sheet("Errors", 2)
        ws_err.cell(row=1, column=1, value="Filename")
        ws_err.cell(row=1, column=2, value="Error")
        for row_idx, (fname, err_msg) in enumerate(errors, 2):
            ws_err.cell(row=row_idx, column=1, value=fname)
            ws_err.cell(row=row_idx, column=2, value=err_msg)
        ws_err.column_dimensions["A"].width = 30
        ws_err.column_dimensions["B"].width = 50

    wb.save(output_path)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Strict namespace-aware ISO 19139 conformance check; produces an Excel report."
    )
    parser.add_argument(
        "input_folder",
        nargs="?",
        default="xml",
        help="Folder containing .xml metadata files (default: xml)",
    )
    args = parser.parse_args()
    folder = Path(args.input_folder)
    folder_name = folder.name
    output_file = "conformance_report_%s.xlsx" % folder_name
    return folder, output_file


def main():
    folder, output_file = parse_args()
    if not folder.exists():
        print("Error: Folder not found: %s" % folder)
        return
    xml_count = len(list(folder.glob("*.xml")))
    if xml_count == 0:
        print("No XML files found in %s" % folder)
        return

    checks = _conformance_checks()
    print("Conformance check (strict ISO 19139 namespaces) for: %s" % folder)
    print("-" * 60)
    print("Found %d XML files" % xml_count)

    results, errors = process_folder(folder, checks)
    if results is None:
        for e in errors:
            print(e)
        return

    for fname in sorted(results.keys()):
        print("Checked: %s" % fname)
    for fname, err in errors:
        print("Skipped: %s — %s" % (fname, err))

    summary = compute_summary(results, checks)
    write_excel_report(results, checks, summary, errors, output_file)
    conformant_count = sum(1 for s in summary if s["Conformant"] == "Yes")
    non_conformant = len(summary) - conformant_count
    print("-" * 60)
    print("Report written: %s" % output_file)
    print("ISO 19139 conformant: %d | Non-conformant (missing mandatory): %d" % (conformant_count, non_conformant))
    if errors:
        print("Skipped (not namespaced ISO 19139): %d" % len(errors))
    print("Done.")


if __name__ == "__main__":
    main()
